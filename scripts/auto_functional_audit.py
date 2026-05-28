"""Automated functional, mathematical and engineering audit.

This script is intentionally broader than ``smoke_app.py`` and narrower than
the full pytest suite.  It exercises the major user-facing workflows with
finite/bounded/trend checks and exits non-zero when a critical check fails.
Heavy tasks remain explicit: the script calls them because the user invoked the
audit, not because Streamlit pages were imported.
"""

from __future__ import annotations

from dataclasses import dataclass
from io import BytesIO
from pathlib import Path
import math
import sys
import zipfile

import numpy as np
import pandas as pd
from openpyxl import load_workbook

sys.path.insert(0, str(Path(__file__).resolve().parents[1]))

import app
from epdm_sim.bayesian_doe import recommend_next_experiment_batch
from epdm_sim.calibration_loop import run_calibration_loop
from epdm_sim.cfd.grid_convergence import run_cfd_grid_convergence
from epdm_sim.cfd.openfoam_export import export_openfoam_case_zip, generate_openfoam_case_files
from epdm_sim.cfd.simple_solver import build_cfd_input_from_flowsheet, run_simple_cfd
from epdm_sim.cfd.visualization import (
    contour_plot,
    export_legacy_vtk,
    mesh_plot,
    reactor_cfd_3d_view,
    streamline_plot,
    surface_plot,
    velocity_vector_plot,
)
from epdm_sim.conservation import run_conservation_checks
from epdm_sim.constrained_window import generate_feasible_windows, rank_process_windows
from epdm_sim.dimensioned import ensure_pressure_Pa, ensure_temperature_K, ensure_viscosity_Pa_s, unit_conversion_trace_dataframe
from epdm_sim.dynamic_residuals import dynamic_residual_acceptance, dynamic_residuals_dataframe
from epdm_sim.dynamic_core.residual_timeseries import dynamic_rhs_residual_acceptance, dynamic_residual_timeseries
from epdm_sim.dynamic_core.rhs_terms import rhs_terms_from_profile
from epdm_sim.dynamic_template_reactor import simulate_template_semibatch_ode
from epdm_sim.engineering_rules import run_all_engineering_rules
from epdm_sim.eos import cubic_eos_details, k_value_comparison
from epdm_sim.equation_binding import run_equation_binding_checks, trend_smoke_results
from epdm_sim.equation_tests import run_equation_code_checks
from epdm_sim.experimental_benchmark import run_experimental_benchmark_checks
from epdm_sim.benchmark_calibration import benchmark_calibration_summary, benchmark_residual_dataframe, recommend_calibration_data_gaps
from epdm_sim.calibrated_property_models import calibrated_property_model_score, calibrated_property_models_dataframe, default_property_model
from epdm_sim.calibrated_property_models import calibrated_property_usage_dataframe
from epdm_sim.benchmark_source_registry import benchmark_source_registry_dataframe, benchmark_source_registry_summary
from epdm_sim.data_lineage import critical_benchmarks_missing_lineage, data_lineage_dataframe, lineage_confidence_score
from epdm_sim.data_lineage_graph import data_lineage_graph_summary
from epdm_sim.dynamic_core.dae_constraints import dae_constraints_dataframe, dae_constraints_status
from epdm_sim.dynamic_core.residual_feedback import dynamic_residual_feedback, residual_feedback_solver_status
from epdm_sim.dynamic_core.solver_decision import dynamic_solver_decision_dataframe
from epdm_sim.dynamic_core.solver_policy import dynamic_solver_policy_dataframe
from epdm_sim.dynamic_core.stability_checks import dynamic_stability_checks_dataframe, dynamic_stability_status
from epdm_sim.dynamic_core.state_invariants import state_invariants_dataframe, state_invariants_status
from epdm_sim.dynamic_core.step_acceptance import dynamic_step_acceptance_dataframe
from epdm_sim.dynamic_core.adaptive_step_control import adaptive_step_control_dataframe, adaptive_step_control_summary
from epdm_sim.dynamic_core.event_detection import dynamic_event_detection_dataframe, event_flags_summary
from epdm_sim.evidence_chain import build_evidence_chain, evidence_gap_dataframe, evidence_weighted_confidence, validate_evidence_chain_completeness
from epdm_sim.evidence_chain_score import critical_evidence_chain_gate, evidence_chain_score_dataframe, evidence_gap_priority_dataframe
from epdm_sim.model_confidence_certificate import confidence_certificate_dataframe, validation_data_upgrade_plan
from epdm_sim.governance_certificate import governance_certificate_dataframe
from epdm_sim.equation_residual_coupling import equation_residual_coupling_dataframe, equation_residual_coupling_summary
from epdm_sim.equation_reverse_check import equation_reverse_check_summary, run_equation_reverse_checks
from epdm_sim.estimation.residual_constrained_fit import run_residual_constrained_fit
from epdm_sim.flash import diagnose_flash_result
from epdm_sim.flowsheet import load_default_config, run_flowsheet
from epdm_sim.io_schema import load_io_schemas
from epdm_sim.digital_twin_3d import build_digital_twin_figure, figure_for_equipment
from epdm_sim.model_audit_report import build_model_audit_report
from epdm_sim.model_confidence_engine import model_confidence_score, recommend_high_value_validation_data
from epdm_sim.model_graph import model_traceability_dataframe, model_traceability_summary
from epdm_sim.model_registry import load_model_registry, validate_model_registry
from epdm_sim.optimizer import optimize_for_grade
from epdm_sim.parameter_estimation import estimate_parameters
from epdm_sim.pareto import generate_pareto_windows
from epdm_sim.posterior import posterior_to_uncertainty_inputs, run_lightweight_mcmc
from epdm_sim.posterior_residual_filter import posterior_residual_filter_dataframe, residual_acceptance_rate
from epdm_sim.preflight import (
    has_blocking_failures,
    run_preflight_for_cfd,
    run_preflight_for_flowsheet,
    run_preflight_for_optimizer,
)
from epdm_sim.property_confidence import propagate_property_uncertainty_to_model_confidence
from epdm_sim.plotting import composition_bar, conversion_bar, flowsheet_block_diagram, sankey_material
from epdm_sim.report import export_excel, export_word_report
from epdm_sim.residual_system import build_flowsheet_residual_system, residual_system_acceptance
from epdm_sim.residual_objective import residual_filter_for_doe, residual_objective_score
from epdm_sim.residual_solver import residual_acceptance_summary, residual_correction_trace_dataframe, residual_solver_dataframe
from epdm_sim.residual_acceptance import doe_residual_acceptance, residual_acceptance_dataframe
from epdm_sim.residual_aware_doe import residual_aware_doe_dataframe
from epdm_sim.residual_aware_optimizer import residual_aware_optimizer_dataframe
from epdm_sim.residual_aware_sampling import residual_aware_sampling_dataframe
from epdm_sim.residual_aware_decision_engine import residual_decision_engine_dataframe
from epdm_sim.ode_diagnostics import ode_diagnostics_dataframe, rhs_terms_diagnostics_dataframe
from epdm_sim.parameter_constraints import parameter_constraint_results_dataframe
from epdm_sim.phase_equilibrium_constraints import phase_equilibrium_constraints_dataframe
from epdm_sim.property_model_selector import property_model_selection_dataframe
from epdm_sim.property_model_bridge import property_model_bridge_dataframe
from epdm_sim.property_model_runtime import property_model_runtime_dataframe
from epdm_sim.property_runtime_context import property_runtime_context_dataframe
from epdm_sim.property_runtime_audit import property_runtime_audit_dataframe
from epdm_sim.industrial_data_package import industrial_data_lineage_dataframe, industrial_data_package_dataframe, load_industrial_data_package, validate_industrial_dataset_schema
from epdm_sim.benchmark_reconciliation import benchmark_reconciliation_dataframe, benchmark_reconciliation_summary
from epdm_sim.repro_package import export_repro_package, load_repro_manifest_from_zip
from epdm_sim.rheology import RheologyParameters, calculate_rheology
from epdm_sim.safety import calculate_safety
from epdm_sim.scaleup import compare_scaleup
from epdm_sim.scientific_benchmarks import run_scientific_benchmarks
from epdm_sim.solubility import liquid_saturation_concentration_mol_L
from epdm_sim.surrogate import (
    predict_with_surrogate,
    surrogate_applicability_warning,
    train_surrogate_from_sensitivity_results,
    validate_surrogate_physics,
)
from epdm_sim.template_config import TemplateProcessConfig
from epdm_sim.template_flowsheet import run_template_flowsheet, template_mass_balance
from epdm_sim.thermo import solve_rachford_rice
from epdm_sim.thermo_consistency import run_thermo_consistency_checks, thermo_physical_constraints_dataframe
from epdm_sim.transport_core import run_transport_core_checks, transport_physical_constraints_dataframe
from epdm_sim.estimation.objectives import residual_aware_parameter_objective
from epdm_sim.solver_core.constrained_solver import constrained_solver_dataframe
from epdm_sim.solver_core.conservation_correction import correction_certificate_dataframe
from epdm_sim.solver_core.conservation_solve_path import conservation_solve_certificate_dataframe
from epdm_sim.solver_core.conservation_jacobian import conservation_jacobian_dataframe
from epdm_sim.solver_core.equation_oriented_solver import equation_oriented_solver_certificate, equation_oriented_solver_gate
from epdm_sim.solver_core.nonlinear_residual_loop import nonlinear_residual_loop_gate, residual_iteration_certificate
from epdm_sim.solver_core.solve_path_integrator import solve_path_integrator_dataframe, solve_path_integrator_gate
from epdm_sim.dynamic_core.adaptive_integrator import adaptive_integrator_dataframe, adaptive_integrator_gate
from epdm_sim.dynamic_core.event_localization import event_localization_dataframe, event_localization_gate
from epdm_sim.solver_core.solver_certificates import solver_certificate_dataframe
from epdm_sim.residual_aware_decision import residual_aware_decision_dataframe
from epdm_sim.data_assimilation import data_assimilation_dataframe, data_assimilation_summary
from epdm_sim.calibration_data_package import calibration_data_lineage_dataframe, calibration_package_dataframe, load_calibration_data_package, validate_calibration_dataset_units
from epdm_sim.ui_audit import run_ui_audit
from epdm_sim.ui_workflow import load_ui_actions, ui_registry_usability_dataframe
from epdm_sim.uncertainty import run_uncertainty_analysis


@dataclass
class AuditCheck:
    """One automated audit result."""

    name: str
    passed: bool
    detail: str
    severity: str = "error"


def _finite(value: object) -> bool:
    """Return whether a scalar or array-like object is finite."""
    try:
        arr = np.asarray(value, dtype=float)
        return bool(np.isfinite(arr).all())
    except Exception:
        return False


def _finite_active_field(result: object, field_name: str) -> bool:
    """Return whether a CFD field is finite inside the active mesh domain."""
    field = result.fields.field(field_name)  # type: ignore[attr-defined]
    mask = result.mesh.mask  # type: ignore[attr-defined]
    return bool(np.isfinite(field[mask]).all())


def _check(checks: list[AuditCheck], name: str, condition: bool, detail: str, severity: str = "error") -> None:
    checks.append(AuditCheck(name, bool(condition), detail, severity))


def _composition_closed(kpis: dict[str, object], tolerance: float = 1.0e-6) -> bool:
    total = float(kpis.get("C2_wt", 0.0)) + float(kpis.get("C3_wt", 0.0)) + float(kpis.get("ENB_wt", 0.0))
    return abs(total - 100.0) <= tolerance


def _copy_config(config: object, **updates: object) -> object:
    """Return a Pydantic config copy without depending on v1/v2 method names."""
    if hasattr(config, "model_copy"):
        return config.model_copy(update=updates)  # type: ignore[attr-defined]
    if hasattr(config, "copy"):
        return config.copy(update=updates)  # type: ignore[attr-defined]
    data = dict(config.__dict__)
    data.update(updates)
    return type(config)(**data)


def run_audit() -> list[AuditCheck]:
    """Run the audit and return all check results."""
    checks: list[AuditCheck] = []

    _check(checks, "app_pages_registered", "数字孪生总览" in app.PAGES and len(app.PAGES) >= 10, f"pages={len(app.PAGES)}")

    modules = load_model_registry()
    registry_errors = validate_model_registry(modules)
    _check(checks, "model_registry_valid", not registry_errors, str(registry_errors))
    schemas = load_io_schemas()
    active_ids = {module.module_id for module in modules if module.status == "active"}
    missing_schema = sorted(active_ids.difference(schemas))
    _check(checks, "active_modules_have_io_schema", not missing_schema, f"missing={missing_schema}")

    actions = load_ui_actions()
    manual_without_task = [action.action_id for action in actions if action.trigger_type == "button_manual" and not action.target_task]
    export_heavy = [
        action.action_id
        for action in actions
        if action.trigger_type == "export" and any(token in action.target_task for token in ("ode", "cfd", "optimization", "posterior"))
    ]
    _check(checks, "ui_actions_manual_have_tasks", not manual_without_task, f"missing_target={manual_without_task}")
    _check(checks, "ui_exports_not_heavy", not export_heavy, f"bad_exports={export_heavy}")
    ui_usability = ui_registry_usability_dataframe(actions)
    failed_usability = ui_usability[~ui_usability["passed"].astype(bool)] if not ui_usability.empty else ui_usability
    _check(
        checks,
        "ui_action_usability_gate",
        failed_usability.empty,
        f"rows={len(ui_usability)}, failed={failed_usability['rule'].tolist() if not failed_usability.empty else []}",
    )
    ui_errors = [issue for issue in run_ui_audit() if issue.severity == "error"]
    _check(checks, "ui_audit_no_errors", not ui_errors, f"errors={len(ui_errors)}")

    cfg = load_default_config()
    preflight = run_preflight_for_flowsheet(cfg)
    _check(checks, "preflight_default_no_blockers", not has_blocking_failures(preflight), f"items={len(preflight)}")
    bad_cfg = _copy_config(cfg, temperature_C=-300.0, pressure_MPa=-1.0, ethylene_kg_h=-2.0)
    bad_preflight = run_preflight_for_flowsheet(bad_cfg)
    _check(checks, "preflight_rejects_illegal_flowsheet", has_blocking_failures(bad_preflight), f"failed={[item.input_name for item in bad_preflight if not item.passed]}")
    cfd_preflight = run_preflight_for_cfd(
        {
            "Nx": 40,
            "Ny": 20,
            "viscosity_Pa_s": 0.02,
            "density_kg_m3": 650.0,
            "Cp_kJ_kgK": 2.2,
            "thermal_conductivity_W_mK": 0.13,
            "diameter_m": 0.08,
            "length_m": 2.0,
            "rpm": cfg.agitation_rpm,
            "heat_generation_W_m3": 1000.0,
        }
    )
    _check(checks, "preflight_cfd_no_blockers", not has_blocking_failures(cfd_preflight), f"items={len(cfd_preflight)}")
    optimizer_preflight = run_preflight_for_optimizer(
        {
            "temperature_C": (80.0, 130.0),
            "pressure_MPa": (0.7, 2.0),
            "enb_kg_h": (1.0, 8.0),
            "hydrogen_g_h": (0.0, 20.0),
        },
        target_grade="Internal_1109_2_commercial_candidate",
    )
    _check(checks, "preflight_optimizer_no_blockers", not has_blocking_failures(optimizer_preflight), f"items={len(optimizer_preflight)}")
    _check(
        checks,
        "dimensioned_input_gate",
        ensure_temperature_K((cfg.temperature_C, "°C")) > 0.0
        and ensure_pressure_Pa((cfg.pressure_MPa, "MPa")) > 0.0
        and ensure_viscosity_Pa_s((1.0, "cP")) > 0.0,
        "temperature/pressure/viscosity unit adapters valid",
    )
    unit_trace = unit_conversion_trace_dataframe()
    _check(checks, "unit_conversion_trace_gate", not unit_trace.empty and (unit_trace["status"] == "ok").all(), f"rows={len(unit_trace)}")

    result = run_flowsheet(cfg)
    kpis = result.kpis
    _check(checks, "flowsheet_polymer_positive", float(kpis["polymer_kg_h"]) > 0.0, f"polymer_kg_h={kpis['polymer_kg_h']:.6g}")
    _check(checks, "flowsheet_template_adapter_used", kpis.get("template_adapter_used") is True, f"template_adapter_used={kpis.get('template_adapter_used')}")
    _check(checks, "product_composition_closes", _composition_closed(kpis), f"C2+C3+ENB={kpis['C2_wt'] + kpis['C3_wt'] + kpis['ENB_wt']:.9f}")
    _check(checks, "heat_duty_positive", float(kpis["heat_duty_kW"]) > 0.0, f"heat_duty_kW={kpis['heat_duty_kW']:.6g}")
    _check(checks, "viscosity_positive", float(kpis["dynamic_viscosity_Pa_s"]) > 0.0, f"mu={kpis['dynamic_viscosity_Pa_s']:.6g}")
    _check(checks, "pressure_drop_positive", float(kpis["pipe_pressure_drop_kPa"]) >= 0.0, f"dP={kpis['pipe_pressure_drop_kPa']:.6g}")
    _check(checks, "all_numeric_kpis_finite", all(_finite(v) for v in kpis.values() if isinstance(v, (int, float))), "numeric KPI finiteness")

    flash_diag_1 = diagnose_flash_result(result.flash1)
    flash_diag_2 = diagnose_flash_result(result.flash2)
    _check(checks, "flash_diagnostics_valid", flash_diag_1.phase_split_valid and flash_diag_2.phase_split_valid, f"vf1={flash_diag_1.vapor_fraction:.6g}, vf2={flash_diag_2.vapor_fraction:.6g}")
    _check(checks, "flash_polymer_nonvolatile", result.flash1.vapor.polymer_mass_kg_h == 0.0 and result.flash2.vapor.polymer_mass_kg_h == 0.0, "polymer vapor mass is zero")

    conservation = run_conservation_checks(result)
    conservation_errors = [item for item in conservation if not item.passed and item.severity == "error"]
    _check(checks, "conservation_no_errors", not conservation_errors, f"errors={[item.balance_type for item in conservation_errors]}")

    rule_results = run_all_engineering_rules(cfg)
    failed_rules = [item.rule_id for item in rule_results if not item.passed]
    _check(checks, "engineering_rules_pass", not failed_rules, f"failed={failed_rules}")

    equation_failures = [item.equation_id for item in run_equation_code_checks() if not item.passed]
    _check(checks, "equation_code_checks_pass", not equation_failures, f"failed={equation_failures}")

    thermo_failures = [item.check_id for item in run_thermo_consistency_checks() if not item.passed and item.severity == "error"]
    _check(checks, "thermo_consistency_no_errors", not thermo_failures, f"failed={thermo_failures}")
    thermo_physical = thermo_physical_constraints_dataframe()
    _check(checks, "thermo_physical_constraints_gate", not thermo_physical.empty and thermo_physical["passed"].all(), f"rows={len(thermo_physical)}")
    phase_constraints = phase_equilibrium_constraints_dataframe()
    _check(checks, "phase_equilibrium_constraints_gate", not phase_constraints.empty and phase_constraints[phase_constraints["severity"] == "error"]["passed"].all(), f"rows={len(phase_constraints)}")
    k_compare = k_value_comparison(["hydrogen", "ethylene", "propylene", "ENB", "hexane", "polymer_EPDM"], 373.15, 1.0e6)
    all_k_values = [value for row in k_compare.values() for value in row.values()]
    _check(checks, "thermo_k_values_positive_finite", all(_finite(value) and float(value) > 0.0 for value in all_k_values), f"components={list(k_compare)}")
    eos_pr = cubic_eos_details("ethylene", 373.15, 1.0e6, eos="PR")
    eos_srk = cubic_eos_details("propylene", 373.15, 1.0e6, eos="SRK")
    eos_values = [
        *eos_pr["Z_roots"],
        eos_pr["phi_v"],
        eos_pr["phi_l"],
        eos_pr["K"],
        *eos_srk["Z_roots"],
        eos_srk["phi_v"],
        eos_srk["phi_l"],
        eos_srk["K"],
    ]
    _check(checks, "eos_z_phi_positive_finite", all(_finite(value) and float(value) > 0.0 for value in eos_values), f"PR_K={eos_pr['K']:.6g}, SRK_K={eos_srk['K']:.6g}")
    rr_vapor = solve_rachford_rice(np.array([0.5, 0.5]), np.array([2.0, 0.5]))
    _check(checks, "rachford_rice_vapor_fraction_bounded", 0.0 <= rr_vapor <= 1.0, f"V={rr_vapor:.6g}")
    henry_low = liquid_saturation_concentration_mol_L("ethylene", "hexane", 373.15, 0.5)
    henry_high = liquid_saturation_concentration_mol_L("ethylene", "hexane", 373.15, 1.5)
    _check(checks, "henry_pressure_monotonic", henry_high >= henry_low >= 0.0 and _finite([henry_low, henry_high]), f"low={henry_low:.6g}, high={henry_high:.6g}")

    safety = calculate_safety(result)
    _check(checks, "thermal_safety_finite", _finite([safety.heat_accumulation_kW, safety.cooling_failure_deltaT_K, safety.MTSR_like_C]), f"risk={safety.runaway_risk_level}")

    scaleup = compare_scaleup(kpis["liquid_density_kg_m3"], kpis["dynamic_viscosity_Pa_s"], rpm=cfg.agitation_rpm)
    _check(checks, "scaleup_metrics_finite", not scaleup.empty and _finite(scaleup.select_dtypes(include="number").to_numpy()), f"rows={len(scaleup)}")

    prop_conf = propagate_property_uncertainty_to_model_confidence()
    _check(checks, "property_confidence_bounded", 0.0 <= float(prop_conf["property_confidence_score"]) <= 100.0, f"score={prop_conf['property_confidence_score']:.6g}")

    rheo_low_shear = calculate_rheology(373.15, kpis["solids_wt"], kpis["Mw"], 1.0, "hexane", RheologyParameters(model="carreau-yasuda"))
    rheo_high_shear = calculate_rheology(373.15, kpis["solids_wt"], kpis["Mw"], 100.0, "hexane", RheologyParameters(model="carreau-yasuda"))
    _check(checks, "rheology_shear_thinning", rheo_high_shear.apparent_viscosity_Pa_s <= rheo_low_shear.apparent_viscosity_Pa_s, f"low={rheo_low_shear.apparent_viscosity_Pa_s:.6g}, high={rheo_high_shear.apparent_viscosity_Pa_s:.6g}")

    figures = {
        "flowsheet_block": flowsheet_block_diagram(),
        "sankey_material": sankey_material(result),
        "conversion_bar": conversion_bar(result),
        "composition_bar": composition_bar(result),
        "digital_twin_overview": build_digital_twin_figure(result, mode="物料流模式", selected_equipment="总览"),
        "reactor_3d": figure_for_equipment("Reactor", result, mode="CFD剖面模式"),
        "flash_3d": figure_for_equipment("Flash2", result),
        "product_3d": figure_for_equipment("Product", result),
    }
    _check(checks, "plotly_figures_generated", all(len(fig.data) > 0 for fig in figures.values()), {name: len(fig.data) for name, fig in figures.items()})

    generic_cfg = TemplateProcessConfig(
        template_id="generic_terpolymerization_apparent",
        monomer_feeds_kg_h={"monomer_A": 2.0, "monomer_B": 2.5, "monomer_C": 1.0},
        solvent_mass_kg_h=20.0,
    )
    generic = run_template_flowsheet(generic_cfg)
    generic_comp = sum(generic.application_kpis["segment_composition_wt"].values())
    generic_mass_error = template_mass_balance(generic)["closure_error_pct"]
    _check(checks, "generic_template_runs", generic.application_kpis["polymer_kg_h"] >= 0.0, f"polymer={generic.application_kpis['polymer_kg_h']:.6g}")
    _check(checks, "generic_template_composition_closes", abs(generic_comp - 100.0) < 1.0e-6 or generic_comp == 0.0, f"sum={generic_comp:.9f}")
    _check(checks, "generic_template_mass_closes", abs(generic_mass_error) < 1.0e-6, f"mass_error_pct={generic_mass_error:.6g}")
    residual_system = build_flowsheet_residual_system(result)
    _check(checks, "residual_system_gate", residual_system.overall_score >= 70.0, f"score={residual_system.overall_score:.6g}")
    residual_acceptance = residual_system_acceptance(residual_system)
    _check(checks, "residual_critical_gate", residual_acceptance["critical_count"] == 0, str(residual_acceptance))
    _check(checks, "residual_objective_gate", residual_objective_score(residual_system) < 50.0, f"penalty={residual_objective_score(residual_system):.6g}")
    residual_solver = residual_solver_dataframe(residual_system)
    residual_corrections = residual_correction_trace_dataframe()
    residual_summary = residual_acceptance_summary(residual_system)
    _check(checks, "residual_solver_gate", bool(residual_summary["passed"]) and not residual_solver.empty, f"summary={residual_summary}")
    _check(checks, "residual_correction_trace_gate", not residual_corrections.empty and residual_corrections["relative_correction_pct"].ge(0.0).all(), f"rows={len(residual_corrections)}")
    equation_solver = equation_oriented_solver_certificate(residual_system)
    equation_solver_gate = equation_oriented_solver_gate(residual_system)
    conservation_jac = conservation_jacobian_dataframe(residual_system)
    _check(checks, "equation_oriented_solver_gate", bool(equation_solver_gate["passed"]) and not equation_solver.empty, f"summary={equation_solver_gate}")
    _check(checks, "conservation_jacobian_gate", not conservation_jac.empty and conservation_jac["finite"].astype(bool).all(), f"rows={len(conservation_jac)}")
    nonlinear_loop = residual_iteration_certificate(residual_system)
    nonlinear_gate = nonlinear_residual_loop_gate(residual_system)
    solve_path = solve_path_integrator_dataframe(residual_system)
    solve_path_gate = solve_path_integrator_gate(residual_system)
    _check(checks, "nonlinear_residual_loop_gate", bool(nonlinear_gate["passed"]) and not nonlinear_loop.empty, f"summary={nonlinear_gate}")
    _check(checks, "solve_path_integrator_gate", bool(solve_path_gate["passed"]) and not solve_path.empty, f"summary={solve_path_gate}")
    residual_acceptance = residual_acceptance_dataframe(residual_system)
    _check(checks, "residual_acceptance_gate", not residual_acceptance.empty and residual_acceptance["passed"].astype(bool).all(), f"rows={len(residual_acceptance)}")
    binding_checks = run_equation_binding_checks()
    _check(checks, "equation_binding_gate", not binding_checks.empty and binding_checks["passed"].all(), f"rows={len(binding_checks)}")
    eq_residual = equation_residual_coupling_dataframe()
    eq_residual_summary = equation_residual_coupling_summary()
    _check(checks, "equation_residual_coupling_gate", bool(eq_residual_summary["passed"]) and not eq_residual.empty, f"summary={eq_residual_summary}")
    trend_checks = trend_smoke_results()
    _check(checks, "dimensional_signature_gate", not trend_checks.empty and trend_checks["passed"].all(), f"rows={len(trend_checks)}")
    benchmark_checks = run_scientific_benchmarks()
    _check(checks, "benchmark_acceptance_gate", not benchmark_checks.empty and benchmark_checks["passed"].all(), f"rows={len(benchmark_checks)}")
    experimental_benchmarks = run_experimental_benchmark_checks()
    _check(checks, "experimental_benchmark_gate", not experimental_benchmarks.empty and experimental_benchmarks[experimental_benchmarks["severity"] == "error"]["passed"].all(), f"rows={len(experimental_benchmarks)}")
    lineage = data_lineage_dataframe()
    lineage_missing = critical_benchmarks_missing_lineage()
    _check(checks, "data_lineage_gate", not lineage.empty and lineage_confidence_score() > 0.0 and lineage_missing["passed"].astype(bool).all(), f"rows={len(lineage)}, score={lineage_confidence_score():.6g}")
    calibration_pkg = load_calibration_data_package()
    calibration_units = validate_calibration_dataset_units(calibration_pkg)
    calibration_lineage = calibration_data_lineage_dataframe(calibration_pkg)
    assimilation = data_assimilation_dataframe()
    assimilation_summary = data_assimilation_summary()
    _check(checks, "calibration_data_package_gate", bool(calibration_units["passed"]) and not calibration_package_dataframe(calibration_pkg).empty and not calibration_lineage.empty, f"units={calibration_units}")
    _check(checks, "data_assimilation_gate", bool(assimilation_summary["passed"]) and not assimilation.empty, f"summary={assimilation_summary}")
    industrial_pkg = load_industrial_data_package()
    industrial_validation = validate_industrial_dataset_schema(industrial_pkg)
    industrial_df = industrial_data_package_dataframe(industrial_pkg)
    industrial_lineage = industrial_data_lineage_dataframe(industrial_pkg)
    reconciliation = benchmark_reconciliation_dataframe(industrial_pkg, {"polymer_mass_closure": 11.5})
    reconciliation_summary = benchmark_reconciliation_summary(industrial_pkg, {"polymer_mass_closure": 11.5})
    _check(
        checks,
        "industrial_data_package_gate",
        bool(industrial_validation["passed"]) and not industrial_df.empty and not industrial_lineage.empty,
        f"validation={industrial_validation}",
    )
    _check(
        checks,
        "benchmark_reconciliation_gate",
        bool(reconciliation_summary["passed"]) and not reconciliation.empty,
        f"summary={reconciliation_summary}",
    )
    lineage_graph_summary = data_lineage_graph_summary()
    _check(checks, "data_lineage_graph_gate", bool(lineage_graph_summary["passed"]), f"summary={lineage_graph_summary}")
    reverse_checks = run_equation_reverse_checks()
    reverse_summary = equation_reverse_check_summary()
    _check(checks, "equation_reverse_check_gate", bool(reverse_summary["passed"]) and not reverse_checks.empty, f"summary={reverse_summary}")
    source_registry = benchmark_source_registry_dataframe()
    source_summary = benchmark_source_registry_summary()
    _check(checks, "benchmark_source_registry_gate", bool(source_summary["passed"]) and not source_registry.empty, f"summary={source_summary}")
    traceability = model_traceability_dataframe()
    traceability_summary = model_traceability_summary()
    _check(checks, "model_traceability_graph_gate", bool(traceability_summary["passed"]) and not traceability.empty, f"summary={traceability_summary}")

    dynamic = simulate_template_semibatch_ode(
        "EPDM_EPM_metallocene_solution",
        config=cfg,
        total_time_min=6.0,
        dt_min=2.0,
        solver_mode="solve_ivp_rk45",
    )
    profile = dynamic.profile
    _check(checks, "dynamic_profile_not_empty", not profile.empty, f"rows={len(profile)}")
    _check(checks, "dynamic_nonnegative", (profile[["polymer_mass_kg", "T_K", "P_Pa"]] >= 0.0).all().all(), "polymer/T/P nonnegative")
    _check(checks, "dynamic_polymer_nondecreasing", profile["polymer_mass_kg"].diff().dropna().ge(-1.0e-10).all(), "polymer monotonic")
    _check(checks, "dynamic_solver_finite", _finite(profile[["T_K", "P_Pa", "polymer_mass_kg"]].to_numpy()), f"solver={dynamic.summary.get('solver_mode_used')}")
    ode_diag = ode_diagnostics_dataframe(dynamic)
    _check(checks, "ode_diagnostics_gate", not ode_diag.empty and ode_diag[ode_diag["severity"] == "error"].empty, f"rows={len(ode_diag)}")
    rhs_diag = rhs_terms_diagnostics_dataframe(dynamic)
    _check(checks, "rhs_diagnostics_gate", not rhs_diag.empty and rhs_diag["finite_check"].all(), f"rows={len(rhs_diag)}")
    dynamic_residuals = dynamic_residuals_dataframe(dynamic)
    dynamic_residual_accept = dynamic_residual_acceptance(dynamic)
    _check(checks, "dynamic_residual_gate", bool(dynamic_residual_accept["passed"]), f"rows={len(dynamic_residuals)}, acceptance={dynamic_residual_accept}")
    rhs_coupling = dynamic_rhs_residual_acceptance(dynamic)
    rhs_profile_terms = rhs_terms_from_profile(dynamic)
    _check(checks, "rhs_residual_coupling_gate", bool(rhs_coupling["passed"]) and not rhs_profile_terms.empty, f"rhs_coupling={rhs_coupling}")
    _check(checks, "dynamic_residual_timeseries_gate", not dynamic_residual_timeseries(dynamic).empty, "dynamic residual time-series available")
    dynamic_feedback = dynamic_residual_feedback(dynamic)
    dynamic_feedback_status = residual_feedback_solver_status(dynamic)
    _check(checks, "dynamic_residual_feedback_gate", bool(dynamic_feedback_status["passed"]) and not dynamic_feedback.empty, f"status={dynamic_feedback_status}")
    dynamic_solver_decision = dynamic_solver_decision_dataframe(dynamic)
    _check(checks, "dynamic_solver_decision_gate", not dynamic_solver_decision.empty and dynamic_solver_decision["residual_acceptance_rate"].between(0, 1).all(), f"rows={len(dynamic_solver_decision)}")
    dynamic_policy = dynamic_solver_policy_dataframe(dynamic)
    _check(checks, "dynamic_solver_policy_gate", not dynamic_policy.empty and dynamic_policy["step_acceptance_rate"].between(0, 1).all(), f"rows={len(dynamic_policy)}")
    dynamic_steps = dynamic_step_acceptance_dataframe(dynamic)
    _check(checks, "dynamic_step_acceptance_gate", not dynamic_steps.empty and dynamic_steps["accepted"].astype(bool).mean() >= 0.5, f"rows={len(dynamic_steps)}")
    adaptive_steps = adaptive_step_control_dataframe(dynamic)
    adaptive_summary = adaptive_step_control_summary(dynamic)
    dynamic_events = dynamic_event_detection_dataframe(dynamic)
    event_summary = event_flags_summary(dynamic)
    _check(checks, "adaptive_step_control_gate", bool(adaptive_summary["passed"]) and not adaptive_steps.empty, f"summary={adaptive_summary}")
    _check(checks, "dynamic_event_detection_gate", bool(event_summary["event_count"] >= 1) and not dynamic_events.empty, f"summary={event_summary}")
    adaptive_integrator = adaptive_integrator_dataframe(dynamic)
    adaptive_integrator_status = adaptive_integrator_gate(dynamic)
    localized_events = event_localization_dataframe(dynamic)
    localized_status = event_localization_gate(dynamic)
    _check(
        checks,
        "adaptive_integrator_gate",
        bool(adaptive_integrator_status["passed"]) and not adaptive_integrator.empty,
        f"summary={adaptive_integrator_status}",
    )
    _check(
        checks,
        "event_localization_gate",
        bool(localized_status["passed"]) and not localized_events.empty,
        f"summary={localized_status}",
    )
    dynamic_stability_checks = dynamic_stability_checks_dataframe(dynamic)
    dynamic_stability_summary = dynamic_stability_status(dynamic)
    _check(checks, "dynamic_stability_checks_gate", bool(dynamic_stability_summary["passed"]) and not dynamic_stability_checks.empty, f"summary={dynamic_stability_summary}")
    dae_constraints = dae_constraints_dataframe(dynamic)
    dae_status = dae_constraints_status(dynamic)
    state_invariants = state_invariants_dataframe(dynamic)
    invariant_status = state_invariants_status(dynamic)
    _check(checks, "dae_state_invariant_gate", bool(dae_status["passed"]) and bool(invariant_status["passed"]) and not state_invariants.empty, f"dae={dae_status}, invariants={invariant_status}")
    explicit_dynamic = simulate_template_semibatch_ode(
        "EPDM_EPM_metallocene_solution",
        config=cfg,
        total_time_min=4.0,
        dt_min=2.0,
        solver_mode="explicit_bounded",
    )
    bdf_dynamic = simulate_template_semibatch_ode(
        "EPDM_EPM_metallocene_solution",
        config=cfg,
        total_time_min=4.0,
        dt_min=2.0,
        solver_mode="solve_ivp_bdf",
    )
    _check(checks, "dynamic_explicit_bounded_runs", not explicit_dynamic.profile.empty and _finite(explicit_dynamic.profile[["T_K", "P_Pa", "polymer_mass_kg"]].to_numpy()), f"rows={len(explicit_dynamic.profile)}")
    _check(checks, "dynamic_bdf_runs_or_fallbacks", not bdf_dynamic.profile.empty and _finite(bdf_dynamic.profile[["T_K", "P_Pa", "polymer_mass_kg"]].to_numpy()), f"solver={bdf_dynamic.summary.get('solver_mode_used')}, fallback={bdf_dynamic.summary.get('fallback_used')}")
    quenched = simulate_template_semibatch_ode(
        "EPDM_EPM_metallocene_solution",
        config=cfg,
        total_time_min=4.0,
        dt_min=1.0,
        solver_mode="explicit_bounded",
    )
    rate_cols = [col for col in quenched.profile.columns if col.startswith("r_")]
    final_rate_sum = float(quenched.profile.loc[quenched.profile.index[-1], rate_cols].abs().sum())
    _check(checks, "dynamic_quench_stops_reaction", final_rate_sum <= 1.0e-8, f"final_rate_sum={final_rate_sum:.6g}")

    generic_dynamic = simulate_template_semibatch_ode(
        "generic_terpolymerization_apparent",
        config=generic_cfg,
        total_time_min=4.0,
        dt_min=2.0,
        solver_mode="explicit_bounded",
    )
    _check(checks, "generic_dynamic_runs", not generic_dynamic.profile.empty, f"rows={len(generic_dynamic.profile)}")

    cfd_input = build_cfd_input_from_flowsheet(result, nx=40, ny=20)
    cfd = run_simple_cfd(cfd_input)
    cfd_diag = cfd.diagnostics
    _check(checks, "cfd_temperature_finite", math.isfinite(cfd_diag.max_temperature_C), f"max_T_C={cfd_diag.max_temperature_C:.6g}")
    _check(checks, "cfd_dead_zone_bounded", 0.0 <= cfd_diag.dead_zone_fraction <= 1.0, f"dead_zone={cfd_diag.dead_zone_fraction:.6g}")
    cfd_fields_ok = all(
        _finite_active_field(cfd, field_name)
        for field_name in ["velocity", "pressure", "temperature", "ENB", "viscosity", "fouling", "wall_shear", "dead_zone_mask", "high_fouling_mask"]
    )
    _check(checks, "cfd_fields_finite", cfd_fields_ok, "velocity/pressure/temperature/scalar/viscosity/fouling finite")
    _check(
        checks,
        "cfd_area_fractions_bounded",
        0.0 <= cfd_diag.high_fouling_zone_area_fraction <= 1.0 and 0.0 <= cfd_diag.low_shear_area_fraction <= 1.0,
        f"high_fouling={cfd_diag.high_fouling_zone_area_fraction:.6g}, low_shear={cfd_diag.low_shear_area_fraction:.6g}",
    )
    _check(checks, "cfd_pump_power_nonnegative", cfd_diag.pump_power_kW >= 0.0, f"pump_power_kW={cfd_diag.pump_power_kW:.6g}")
    cfd_figures = {
        "mesh": mesh_plot(cfd),
        "contour_temperature": contour_plot(cfd, "temperature"),
        "vector": velocity_vector_plot(cfd),
        "streamline": streamline_plot(cfd),
        "surface_viscosity": surface_plot(cfd, "viscosity"),
        "reactor_3d_cfd": reactor_cfd_3d_view(cfd, "temperature"),
    }
    _check(checks, "cfd_plotly_figures_generated", all(len(fig.data) > 0 for fig in cfd_figures.values()), {name: len(fig.data) for name, fig in cfd_figures.items()})
    vtk_text = export_legacy_vtk(cfd).decode("utf-8", errors="ignore")
    _check(checks, "cfd_vtk_contains_masks", all(token in vtk_text for token in ["dead_zone_mask", "high_fouling_mask", "wall_shear"]), "VTK mask/scalar names present")
    openfoam_files = generate_openfoam_case_files(cfd_input)
    required_openfoam = {"system/blockMeshDict", "0/U", "0/p", "0/T", "constant/transportProperties", "system/controlDict", "system/fvSchemes", "system/fvSolution"}
    _check(checks, "openfoam_case_files_present", required_openfoam.issubset(openfoam_files), f"missing={sorted(required_openfoam.difference(openfoam_files))}")
    with zipfile.ZipFile(BytesIO(export_openfoam_case_zip(cfd_input)), "r") as openfoam_zip:
        openfoam_zip_names = set(openfoam_zip.namelist())
    _check(checks, "openfoam_zip_contains_required_files", required_openfoam.issubset(openfoam_zip_names), f"missing={sorted(required_openfoam.difference(openfoam_zip_names))}")
    grid = run_cfd_grid_convergence(cfd_input, grids=[(30, 15), (40, 20)])
    _check(checks, "cfd_grid_convergence_finite", _finite(grid.metrics.select_dtypes(include="number").to_numpy()), f"score={grid.convergence_score:.6g}")
    transport_checks = run_transport_core_checks()
    _check(checks, "transport_core_gate", not transport_checks.empty and transport_checks["passed"].all(), f"rows={len(transport_checks)}")
    transport_physical = transport_physical_constraints_dataframe()
    _check(checks, "transport_physical_constraints_gate", not transport_physical.empty and transport_physical["passed"].all(), f"rows={len(transport_physical)}")

    doe = recommend_next_experiment_batch(cfg, n=3, seed=3)
    _check(checks, "bayesian_doe_candidates", len(doe) >= 1, f"rows={len(doe)}")
    _check(checks, "residual_aware_doe_gate", residual_filter_for_doe({"residual_system": residual_system})["passed"] and doe_residual_acceptance({"residual_system": residual_system})["passed"], "default residual system accepted for DOE")
    residual_doe = residual_aware_doe_dataframe(residual_system)
    _check(checks, "residual_aware_doe_v6_2_gate", not residual_doe.empty and not residual_doe["rejected"].astype(bool).any(), f"rows={len(residual_doe)}")
    if not doe.empty:
        risk_col = "predicted_risk" if "predicted_risk" in doe.columns else None
        if risk_col:
            _check(checks, "bayesian_doe_risk_bounded", doe[risk_col].between(0, 1).all(), "risk in [0,1]")

    uncertainty = run_uncertainty_analysis(cfg, n_samples=6, seed=5)
    _check(checks, "uncertainty_intervals_finite", not uncertainty.confidence_intervals.empty and _finite(uncertainty.confidence_intervals.select_dtypes(include="number").to_numpy()), f"rows={len(uncertainty.confidence_intervals)}")
    _check(checks, "uncertainty_risk_bounded", all(0.0 <= float(value) <= 1.0 for value in uncertainty.risk_probabilities.values()), str(uncertainty.risk_probabilities))
    parameter_constraints = parameter_constraint_results_dataframe({"k_E_ref": 100.0, "Mw0": 350000.0, "ktr_H2": 0.2})
    _check(checks, "parameter_constraints_gate", parameter_constraints["passed"].all(), f"rows={len(parameter_constraints)}")

    posterior = run_lightweight_mcmc(n_steps=20, seed=13)
    _check(checks, "posterior_finite", not posterior.samples.empty and _finite(posterior.samples.to_numpy()), f"acceptance={posterior.acceptance_rate:.6g}")
    posterior_uncertainty = posterior_to_uncertainty_inputs(posterior)
    _check(checks, "posterior_uncertainty_finite", posterior_uncertainty and all(_finite(v) for v in posterior_uncertainty.values()), f"params={len(posterior_uncertainty)}")
    _check(checks, "posterior_residual_acceptance_gate", residual_aware_parameter_objective(1.0, residual_system) < 10.0, "posterior/calibration objective accepts default residual system")
    posterior_filter = posterior_residual_filter_dataframe(posterior.samples, residual_system)
    _check(checks, "posterior_residual_filter_gate", 0.0 <= residual_acceptance_rate(posterior.samples, residual_system) <= 1.0 and not posterior_filter.empty, f"rate={residual_acceptance_rate(posterior.samples, residual_system):.6g}")
    residual_decision = residual_aware_decision_dataframe(residual_system)
    _check(checks, "residual_aware_decision_gate", not residual_decision.empty and residual_decision["risk_probability"].between(0, 1).all(), f"rows={len(residual_decision)}")
    residual_sampling = residual_aware_sampling_dataframe(result_or_system=residual_system)
    _check(checks, "residual_aware_sampling_gate", not residual_sampling.empty and residual_sampling["uncertainty_risk_probability"].between(0, 1).all() and not residual_sampling["rejected"].astype(bool).any(), f"rows={len(residual_sampling)}")
    residual_engine = residual_decision_engine_dataframe(result_or_system=residual_system)
    _check(
        checks,
        "residual_decision_engine_gate",
        not residual_engine.empty
        and residual_engine["uncertainty_risk_probability"].between(0, 1).all()
        and not residual_engine["rejected"].astype(bool).any(),
        f"rows={len(residual_engine)}",
    )

    calibration_loop = run_calibration_loop(config=cfg, target_metrics=["C2_wt", "ENB_wt", "Mw"])
    _check(checks, "calibration_loop_recommends_experiments", not calibration_loop.recommended_experiments.empty, f"rows={len(calibration_loop.recommended_experiments)}")
    residual_fit = run_residual_constrained_fit(
        initial_params={"k_E_ref": 100.0, "Mw0": 350000.0, "ktr_H2": 0.2},
        result_or_residual_system=residual_system,
        target_units={"C2_wt": "wt%", "Mw": "g/mol"},
        data_residual=1.0,
    )
    _check(checks, "residual_constrained_fit_gate", residual_fit.accepted and _finite(residual_fit.objective), f"objective={residual_fit.objective:.6g}")
    benchmark_cal = benchmark_calibration_summary(kpis)
    benchmark_residuals = benchmark_residual_dataframe(kpis)
    data_gaps = recommend_calibration_data_gaps(kpis)
    _check(checks, "benchmark_calibration_gate", not benchmark_cal.empty and _finite(benchmark_cal.select_dtypes(include="number").to_numpy()), f"rows={len(benchmark_cal)}")
    _check(checks, "benchmark_data_gap_gate", not data_gaps.empty and data_gaps["recommended_action"].astype(str).str.len().gt(0).all(), f"rows={len(data_gaps)}")
    property_models = calibrated_property_models_dataframe([default_property_model()])
    property_usage = calibrated_property_usage_dataframe([default_property_model()])
    _check(checks, "calibrated_property_model_gate", not property_models.empty and 0.0 <= calibrated_property_model_score([default_property_model()]) <= 100.0, f"score={calibrated_property_model_score([default_property_model()]):.6g}")
    _check(checks, "calibrated_property_usage_gate", not property_usage.empty and property_usage["passed"].astype(bool).all(), f"rows={len(property_usage)}")
    property_selection = property_model_selection_dataframe()
    _check(checks, "property_model_selector_gate", not property_selection.empty and property_selection["confidence_score"].between(0, 100).all(), f"rows={len(property_selection)}")
    property_bridge = property_model_bridge_dataframe()
    _check(checks, "property_model_bridge_gate", not property_bridge.empty and property_bridge["passed"].astype(bool).all(), f"rows={len(property_bridge)}")
    property_runtime = property_model_runtime_dataframe(conditions={"temperature_C": cfg.temperature_C, "pressure_MPa": cfg.pressure_MPa, "solids_wt": kpis["solids_wt"]})
    _check(checks, "property_model_runtime_gate", not property_runtime.empty and property_runtime["passed"].astype(bool).all(), f"rows={len(property_runtime)}")
    property_context = property_runtime_context_dataframe(result, conditions={"temperature_C": cfg.temperature_C, "pressure_MPa": cfg.pressure_MPa, "solids_wt": kpis["solids_wt"]})
    _check(checks, "property_runtime_context_gate", not property_context.empty and property_context["passed"].astype(bool).all(), f"rows={len(property_context)}")
    property_runtime_audit = property_runtime_audit_dataframe(result, conditions={"temperature_C": cfg.temperature_C, "pressure_MPa": cfg.pressure_MPa, "solids_wt": kpis["solids_wt"]})
    _check(checks, "property_runtime_audit_gate", not property_runtime_audit.empty and property_runtime_audit["passed"].astype(bool).all(), f"rows={len(property_runtime_audit)}")
    param_est = estimate_parameters(target="C2_wt", method="least_squares", max_nfev=3, model_mode="empirical_proxy")
    _check(
        checks,
        "parameter_estimation_small_budget_runs",
        bool(param_est.fitted_params) and _finite(list(param_est.fitted_params.values())) and _finite(param_est.fitting_runtime_s),
        f"params={len(param_est.fitted_params)}, runtime={param_est.fitting_runtime_s:.6g}, failures={len(param_est.run_failures)}",
    )
    opt = optimize_for_grade(cfg, grade_id="Internal_1109_2_commercial_candidate", maxiter=1)
    _check(checks, "optimizer_small_budget_runs", _finite(opt.objective) and bool(opt.kpis), f"objective={opt.objective:.6g}, success={opt.success}")
    pareto = generate_pareto_windows(cfg, grade_id="Internal_1109_2_commercial_candidate", n_samples=6, seed=7)
    _check(checks, "pareto_small_sample_runs", not pareto.candidates.empty and _finite(pareto.candidates.select_dtypes(include="number").to_numpy()), f"candidates={len(pareto.candidates)}, frontier={len(pareto.frontier)}")

    windows = rank_process_windows(generate_feasible_windows(cfg))
    _check(checks, "constrained_windows_feasible", bool(windows), f"windows={len(windows)}")
    if windows:
        margins_ok = all(all(float(value) >= 0.0 for value in window.constraint_margins.values()) for window in windows)
        _check(checks, "constrained_windows_margins_nonnegative", margins_ok, f"first={windows[0].constraint_margins}")

    surrogate_training = pd.DataFrame(
        [
            {"temperature_C": 90.0, "solids_wt": 8.0, "hydrogen_g_h": 2.0, "Mw": 460000.0, "viscosity": 0.010},
            {"temperature_C": 100.0, "solids_wt": 10.0, "hydrogen_g_h": 5.0, "Mw": 360000.0, "viscosity": 0.014},
            {"temperature_C": 110.0, "solids_wt": 14.0, "hydrogen_g_h": 9.0, "Mw": 280000.0, "viscosity": 0.022},
            {"temperature_C": 120.0, "solids_wt": 18.0, "hydrogen_g_h": 12.0, "Mw": 230000.0, "viscosity": 0.034},
        ]
    )
    surrogate = train_surrogate_from_sensitivity_results(surrogate_training, "viscosity", ["temperature_C", "solids_wt"])
    surrogate_checks = validate_surrogate_physics(surrogate)
    surrogate_pred = predict_with_surrogate(surrogate, {"temperature_C": 100.0, "solids_wt": 12.0})
    _check(checks, "surrogate_physics_pass", surrogate_checks["passed"].all(), surrogate_checks.to_dict(orient="records"))
    _check(checks, "surrogate_prediction_finite", _finite(surrogate_pred), f"pred={surrogate_pred.tolist()}")
    _check(checks, "surrogate_out_of_range_warns", bool(surrogate_applicability_warning(surrogate, {"temperature_C": 150.0, "solids_wt": 12.0})), "out-of-range warning exists")

    audit = build_model_audit_report(result, uncertainty=uncertainty)
    audit_score = float(audit.model_confidence_card.overall_score)
    _check(checks, "model_audit_score_bounded", 0.0 <= audit_score <= 100.0, f"score={audit_score:.6g}")
    confidence = model_confidence_score(residual_system=residual_system, model_outputs=kpis)
    _check(checks, "model_confidence_engine_gate", bool(confidence["passed"]) and 0.0 <= float(confidence["overall_score"]) <= 100.0, f"score={confidence['overall_score']:.6g}")
    _check(checks, "validation_data_recommendations_gate", not recommend_high_value_validation_data().empty, "validation data gap recommendations available")
    evidence_chain = build_evidence_chain()
    evidence_status = validate_evidence_chain_completeness(evidence_chain)
    evidence_confidence = evidence_weighted_confidence(evidence_chain)
    _check(checks, "evidence_chain_gate", bool(evidence_status["passed"]) and bool(evidence_confidence["passed"]) and not evidence_gap_dataframe().empty, f"status={evidence_status}, confidence={evidence_confidence}")
    evidence_score = critical_evidence_chain_gate(evidence_chain)
    _check(checks, "evidence_chain_score_gate", bool(evidence_score["passed"]) and not evidence_gap_priority_dataframe(evidence_chain).empty and not evidence_chain_score_dataframe(evidence_chain).empty, f"score={evidence_score}")
    confidence_certificate = confidence_certificate_dataframe(residual_system=residual_system, model_outputs=kpis)
    _check(checks, "confidence_certificate_gate", not confidence_certificate.empty and confidence_certificate["passed"].astype(bool).all() and not validation_data_upgrade_plan().empty, f"rows={len(confidence_certificate)}")
    governance_certificate = governance_certificate_dataframe(result)
    _check(checks, "governance_certificate_gate", not governance_certificate.empty and governance_certificate["passed"].astype(bool).all(), f"rows={len(governance_certificate)}")
    constrained_status = constrained_solver_dataframe(residual_system)
    solver_certificate = solver_certificate_dataframe(residual_system)
    correction_certificates = correction_certificate_dataframe(residual_system)
    conservation_solve = conservation_solve_certificate_dataframe(residual_system)
    _check(checks, "constrained_solver_gate", bool(constrained_status["accepted"].iloc[0]) and not constrained_status.empty, f"rows={len(constrained_status)}")
    _check(checks, "solver_certificate_gate", bool(solver_certificate["solver_certificate_passed"].iloc[0]), f"rows={len(solver_certificate)}")
    _check(checks, "conservation_correction_gate", not correction_certificates.empty and not correction_certificates["rejected"].astype(bool).any(), f"rows={len(correction_certificates)}")
    _check(checks, "conservation_solve_path_gate", not conservation_solve.empty and not conservation_solve["rejected"].astype(bool).any(), f"rows={len(conservation_solve)}")
    residual_optimizer = residual_aware_optimizer_dataframe(residual_system)
    _check(checks, "residual_aware_optimizer_gate", not residual_optimizer.empty and not residual_optimizer["rejected"].astype(bool).any(), f"rows={len(residual_optimizer)}")

    xlsx = export_excel(result)
    docx = export_word_report(result)
    _check(checks, "excel_export_nonempty", len(xlsx) > 1000, f"bytes={len(xlsx)}")
    _check(checks, "word_export_nonempty", len(docx) > 1000, f"bytes={len(docx)}")
    workbook = load_workbook(BytesIO(xlsx), read_only=True)
    _check(checks, "excel_sheet_name_compatibility_gate", all(len(name) <= 31 for name in workbook.sheetnames), f"long={[name for name in workbook.sheetnames if len(name) > 31]}")
    required_sheets = {
        "stream table",
        "template_config",
        "template_flowsheet",
        "template_kpis",
        "equation_code_checks",
        "model_registry_snapshot",
        "audit_trail",
        "preflight",
        "conservation",
        "thermo_consistency",
        "property_confidence",
        "model_audit",
        "conservation_correction",
        "conservation_solve_path",
        "conservation_solve_cert",
        "equation_oriented_solver",
        "conservation_jacobian",
        "nonlinear_residual_loop",
        "solve_path_integrator",
        "calibration_data_package",
        "data_assimilation",
        "industrial_data_package",
        "benchmark_reconciliation",
        "property_model_bridge",
        "property_model_runtime",
        "property_runtime_context",
        "property_runtime_audit",
        "dynamic_solver_decision",
        "dynamic_solver_policy",
        "dynamic_step_acceptance",
        "adaptive_step_control",
        "dynamic_event_detection",
        "adaptive_integrator",
        "event_localization",
        "residual_aware_decision",
        "residual_decision_engine",
        "residual_aware_optimizer",
        "residual_aware_sampling",
        "evidence_chain",
        "evidence_chain_score",
        "evidence_gap_priority",
        "confidence_certificate",
        "governance_certificate",
        "validation_upgrade_plan",
        "V6_2_audit_summary",
        "V6_3_audit_summary",
        "V6_4_audit_summary",
    }
    _check(checks, "excel_required_sheets", required_sheets.issubset(set(workbook.sheetnames)), f"missing={sorted(required_sheets.difference(workbook.sheetnames))}")

    package = export_repro_package(result, report_xlsx=xlsx, report_docx=docx, test_status="auto_functional_audit")
    manifest = load_repro_manifest_from_zip(package)
    with zipfile.ZipFile(BytesIO(package), "r") as zf:
        names = set(zf.namelist())
    _check(checks, "repro_manifest_version", str(manifest.get("app_version", "")).startswith("V6.4"), f"version={manifest.get('app_version')}")
    _check(checks, "repro_contains_audit_trail", "audit_trail.csv" in names, f"files={len(names)}")
    v60_repro_files = {
        "data_lineage.csv",
        "data_lineage_graph.csv",
        "benchmark_sources.csv",
        "benchmark_lineage.csv",
        "calibrated_property_models.csv",
        "calibrated_property_usage.csv",
        "property_model_selection.csv",
        "equation_reverse_check.csv",
        "equation_residual_coupling.csv",
        "residual_acceptance.csv",
        "model_traceability_graph.csv",
        "equation_graph.csv",
        "residual_graph.csv",
        "validation_evidence.csv",
        "confidence_decomposition.csv",
        "solver_certificate.csv",
        "correction_certificates.csv",
        "property_model_bridge.csv",
        "property_model_runtime.csv",
        "dynamic_solver_decision.csv",
        "dynamic_solver_policy.csv",
        "dynamic_step_acceptance.csv",
        "residual_aware_decision.csv",
        "residual_aware_optimizer.csv",
        "residual_aware_doe.csv",
        "evidence_chain.csv",
        "evidence_gaps.csv",
        "evidence_chain_score.csv",
        "evidence_gap_priority.csv",
        "conservation_solve_certificate.csv",
        "equation_oriented_solver.csv",
        "conservation_jacobian.csv",
        "nonlinear_residual_loop.csv",
        "solve_path_integrator.csv",
        "calibration_data_package.csv",
        "calibration_data_lineage.csv",
        "data_assimilation.csv",
        "industrial_data_package.csv",
        "industrial_data_lineage.csv",
        "benchmark_reconciliation.csv",
        "property_runtime_context.csv",
        "property_runtime_audit.csv",
        "adaptive_step_control.csv",
        "dynamic_event_detection.csv",
        "adaptive_integrator.csv",
        "event_localization.csv",
        "residual_aware_sampling.csv",
        "residual_decision_engine.csv",
        "confidence_certificate.csv",
        "governance_certificate.csv",
        "validation_upgrade_plan.csv",
    }
    _check(checks, "report_repro_industrial_audit_gate", v60_repro_files.issubset(names), f"missing={sorted(v60_repro_files.difference(names))}")

    return checks


def main() -> int:
    checks = run_audit()
    rows = [check.__dict__ for check in checks]
    out_dir = Path("tmp_smoke_outputs")
    out_dir.mkdir(exist_ok=True)
    pd.DataFrame(rows).to_csv(out_dir / "auto_functional_audit.csv", index=False, encoding="utf-8-sig")
    failed = [check for check in checks if not check.passed and check.severity == "error"]
    print(f"auto functional audit: {len(checks) - len(failed)}/{len(checks)} checks passed")
    for check in checks:
        status = "PASS" if check.passed else check.severity.upper()
        print(f"[{status}] {check.name}: {check.detail}")
    return 1 if failed else 0


if __name__ == "__main__":
    raise SystemExit(main())
