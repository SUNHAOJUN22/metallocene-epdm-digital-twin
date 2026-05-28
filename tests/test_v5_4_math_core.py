from __future__ import annotations

from io import BytesIO
import zipfile

import numpy as np
from openpyxl import load_workbook

from epdm_sim.dimensioned import unit_conversion_trace_dataframe
from epdm_sim.dynamic_residuals import DynamicResidualPoint, dynamic_residual_acceptance, dynamic_residual_system, dynamic_residuals_dataframe
from epdm_sim.dynamic_template_reactor import simulate_template_semibatch_ode
from epdm_sim.experimental_benchmark import (
    benchmark_confidence_score,
    benchmark_data_hash,
    experimental_benchmarks_dataframe,
    load_experimental_benchmarks,
    run_experimental_benchmark_checks,
)
from epdm_sim.flash import Flash
from epdm_sim.flowsheet import load_default_config, run_flowsheet
from epdm_sim.heat_balance import calculate_heat_balance
from epdm_sim.kinetics import KineticParameters
from epdm_sim.phase_equilibrium_constraints import (
    PhaseEquilibriumConstraint,
    classify_z_roots,
    flash_residuals_dataframe,
    k_value_ordering_dataframe,
    phase_equilibrium_constraints_dataframe,
)
from epdm_sim.repro_package import export_repro_package, load_repro_manifest_from_zip
from epdm_sim.report import export_excel
from epdm_sim.residual_objective import (
    reject_if_critical_residual,
    residual_diagnostics_dataframe,
    residual_filter_for_doe,
    residual_objective_score,
    residual_penalty_for_optimizer,
)
from epdm_sim.residual_system import ResidualSystem, build_flowsheet_residual_system, make_residual
from epdm_sim.rheology import apparent_viscosity_from_zero_shear, calculate_rheology
from epdm_sim.state_vector import build_state_layout_from_template
from epdm_sim.template_ode_rhs import build_template_ode_context, initial_template_ode_state


def test_unit_safe_entries_and_conversion_trace():
    cfg = load_default_config()
    result = run_flowsheet(cfg)
    inlet = result.streams["Quenched solution"]
    flash = Flash("unit-safe")
    by_engineering_units = flash.calculate(inlet, (100.0, "°C"), (0.1, "MPa"))
    by_base_units = flash.calculate(inlet, 373.15, 100000.0)
    assert abs(by_engineering_units.vapor_fraction - by_base_units.vapor_fraction) < 1.0e-12

    heat_kjh = calculate_heat_balance({"ethylene": 1.0}, 10.0, 2.0, preheat_kJ_h=(3600.0, "kJ/h"))
    heat_kw = calculate_heat_balance({"ethylene": 1.0}, 10.0, 2.0, preheat_kJ_h=(1.0, "kW"))
    assert abs(heat_kjh.preheat_kW - heat_kw.preheat_kW) < 1.0e-12

    rheo_c = calculate_rheology((100.0, "°C"), 12.0, 300000.0, 10.0)
    rheo_k = calculate_rheology(373.15, 12.0, 300000.0, 10.0)
    assert abs(rheo_c.apparent_viscosity_Pa_s - rheo_k.apparent_viscosity_Pa_s) < 1.0e-12
    assert apparent_viscosity_from_zero_shear((1.0, "cP"), 10.0) > 0.0

    layout = build_state_layout_from_template("EPDM_EPM_metallocene_solution")
    context = build_template_ode_context("EPDM_EPM_metallocene_solution", layout, KineticParameters(), cfg, total_time_min=(180.0, "s"))
    state = initial_template_ode_state(
        context,
        solvent_mass_kg=1.0,
        temperature_K=(100.0, "°C"),
        pressure_Pa=(1.0, "MPa"),
        catalyst_active_mol=1.0e-6,
    )
    assert state["T_K"] > 273.15
    assert state["P_Pa"] == 1.0e6

    trace = unit_conversion_trace_dataframe()
    assert not trace.empty
    assert set(["field", "unit", "si_unit", "status"]).issubset(trace.columns)
    assert (trace["status"] == "ok").all()


def test_residual_objective_rejects_critical_residuals():
    result = run_flowsheet(load_default_config())
    default_system = build_flowsheet_residual_system(result)
    assert residual_objective_score(default_system) < 50.0
    assert residual_penalty_for_optimizer(default_system) < 50.0
    assert not reject_if_critical_residual(default_system)["rejected"]
    assert residual_filter_for_doe({"residual_system": default_system})["passed"]
    assert not residual_diagnostics_dataframe(default_system).empty

    critical = make_residual("forced_polymer_vapor", "polymer vapor = 0", 1.0, 0.0, "kg/h", 1.0e-12, "flash", "Keep polymer in liquid.", "critical")
    bad_system = ResidualSystem(phase_residuals=[critical], overall_score=0.0)
    assert residual_objective_score(bad_system) >= 1000.0
    assert reject_if_critical_residual(bad_system)["rejected"]
    assert not residual_filter_for_doe({"residual_system": bad_system})["passed"]


def test_dynamic_residuals_are_finite_and_quench_is_checked():
    dynamic = simulate_template_semibatch_ode(total_time_min=4.0, dt_min=1.0, solver_mode="explicit_bounded")
    point = DynamicResidualPoint(0.0, "example", 0.0, "-", "ok", True, "example physical meaning")
    assert point.as_dict()["passed"] is True
    df = dynamic_residuals_dataframe(dynamic)
    assert not df.empty
    assert np.isfinite(df["value"].to_numpy(dtype=float)).all()
    assert dynamic_residual_acceptance(dynamic)["passed"]
    system = dynamic_residual_system(dynamic)
    assert system.overall_score >= 70.0


def test_phase_equilibrium_constraints_and_flash_residuals():
    marker = PhaseEquilibriumConstraint("marker", True, "ok", 1.0, "-", "marker")
    assert marker.as_dict()["passed"] is True
    root = classify_z_roots("ethylene", 373.15, 1.0e6, "PR")
    assert root["root_order_valid"]
    assert root["phi_v"] > 0.0
    k_df = k_value_ordering_dataframe()
    assert not k_df.empty
    assert (k_df["K"] > 0.0).all()
    result = run_flowsheet(load_default_config())
    flash_df = flash_residuals_dataframe(result.flash1)
    assert not flash_df.empty
    assert flash_df.loc[flash_df["residual_id"] == "flash_polymer_vapor", "passed"].iloc[0]
    constraints = phase_equilibrium_constraints_dataframe(result)
    assert not constraints.empty
    assert constraints[constraints["severity"] == "error"]["passed"].all()


def test_experimental_benchmarks_metadata_and_confidence():
    records = load_experimental_benchmarks()
    assert len(records) >= 3
    assert benchmark_data_hash(records[0])
    df = experimental_benchmarks_dataframe()
    assert not df.empty
    assert df["hash_matches"].all()
    checks = run_experimental_benchmark_checks()
    assert not checks.empty
    assert checks[checks["severity"] == "error"]["passed"].all()
    assert 0.0 <= benchmark_confidence_score() <= 100.0


def test_report_and_repro_package_include_v5_4_math_core_artifacts():
    result = run_flowsheet(load_default_config())
    xlsx = export_excel(result)
    workbook = load_workbook(BytesIO(xlsx), read_only=True)
    required = {
        "unit_conversion_trace",
        "residual_objective",
        "dynamic_residuals",
        "phase_equilibrium_constraints",
        "experimental_benchmarks",
        "residual_aware_optimization",
        "residual_aware_doe",
    }
    assert required.issubset(set(workbook.sheetnames))
    package = export_repro_package(result, report_xlsx=xlsx, test_status="v5_4_test")
    manifest = load_repro_manifest_from_zip(package)
    assert manifest["app_version"].startswith("V6.4")
    with zipfile.ZipFile(BytesIO(package)) as zf:
        names = set(zf.namelist())
    assert {"experimental_benchmarks.json", "residual_system.csv", "benchmark_snapshot.json"}.issubset(names)

