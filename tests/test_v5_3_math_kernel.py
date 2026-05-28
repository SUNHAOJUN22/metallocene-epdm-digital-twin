import io
import zipfile

import pytest
from openpyxl import load_workbook

from epdm_sim.bayesian_doe import recommend_next_experiment_batch
from epdm_sim.constrained_window import generate_feasible_windows
from epdm_sim.dimensioned import (
    DimensionedValue,
    ensure_concentration_mol_L,
    ensure_density_kg_m3,
    ensure_length_m,
    ensure_mass_flow_kg_h,
    ensure_molar_flow_mol_h,
    ensure_power_kW,
    ensure_pressure_Pa,
    ensure_temperature_K,
    ensure_time_min,
    ensure_viscosity_Pa_s,
)
from epdm_sim.equation_binding import equation_binding_dataframe, validate_equation_bindings
from epdm_sim.flowsheet import load_default_config, run_flowsheet
from epdm_sim.ode_diagnostics import RHSTermDiagnostic, rhs_terms_diagnostics_dataframe
from epdm_sim.report import export_excel
from epdm_sim.repro_package import export_repro_package, load_repro_manifest_from_zip
from epdm_sim.residual_system import (
    build_flowsheet_residual_system,
    critical_residuals,
    make_residual,
    residual_system_acceptance,
)
from epdm_sim.scientific_benchmarks import benchmark_definitions, run_scientific_benchmarks
from epdm_sim.thermo_consistency import thermo_physical_constraints_dataframe
from epdm_sim.transport_core import transport_physical_constraints_dataframe


def test_v5_3_dimensioned_adapters_are_explicit_and_bounded():
    assert ensure_temperature_K((100.0, "°C")) == pytest.approx(373.15)
    assert ensure_pressure_Pa((1.2, "MPa")) == pytest.approx(1.2e6)
    assert ensure_mass_flow_kg_h((1500.0, "g/h")) == pytest.approx(1.5)
    assert ensure_molar_flow_mol_h((2.0, "kmol/h")) == pytest.approx(2000.0)
    assert ensure_concentration_mol_L((1000.0, "mol/m3")) == pytest.approx(1.0)
    assert ensure_power_kW((3600.0, "kJ/h")) == pytest.approx(1.0)
    assert ensure_viscosity_Pa_s((1.0, "cP")) == pytest.approx(0.001)
    assert ensure_density_kg_m3(DimensionedValue(720.0, "kg/m3")) == pytest.approx(720.0)
    assert ensure_length_m((50.0, "mm")) == pytest.approx(0.05)
    assert ensure_time_min((3600.0, "s")) == pytest.approx(60.0)
    with pytest.raises(ValueError):
        ensure_pressure_Pa((100.0, "°C"))
    with pytest.raises(ValueError):
        ensure_temperature_K((-300.0, "°C"))


def test_v5_3_residual_system_blocks_critical_failures():
    result = run_flowsheet(load_default_config())
    system = build_flowsheet_residual_system(result)
    accepted = residual_system_acceptance(system)
    assert accepted["passed"]
    assert accepted["critical_count"] == 0
    critical = make_residual(
        "polymer_vapor_forced",
        "polymer vapor = 0",
        1.0,
        0.0,
        "kg/h",
        1e-12,
        "flash",
        "force polymer pseudo to liquid",
        "critical",
    )
    from epdm_sim.residual_system import ResidualSystem

    failed = ResidualSystem(phase_residuals=[critical], overall_score=50.0)
    assert critical_residuals(failed)
    assert not residual_system_acceptance(failed)["passed"]


def test_v5_3_equation_binding_and_benchmarks_are_executable_specs():
    errors = validate_equation_bindings()
    assert errors == []
    bindings = equation_binding_dataframe()
    critical = bindings[bindings["implementation_function"].astype(str).str.len() > 0]
    assert not critical.empty
    assert critical["importable"].all()
    assert critical["benchmark_id"].astype(str).str.len().gt(0).all()
    assert critical["residual_id"].astype(str).str.len().gt(0).all()
    definitions = benchmark_definitions()
    assert not definitions.empty
    assert definitions["model_version"].astype(str).str.startswith("V6.4").all()
    assert definitions["residual_id"].astype(str).str.len().gt(0).all()
    assert run_scientific_benchmarks()["passed"].all()


def test_v5_3_rhs_thermo_transport_constraints_are_reportable():
    assert RHSTermDiagnostic("feed_term", 0.0, "mol/min", "liquid_moles", "feed", True, True).finite_check
    rhs = rhs_terms_diagnostics_dataframe()
    assert {"term", "value", "unit", "affected_state", "physical_meaning", "finite_check"}.issubset(rhs.columns)
    assert rhs["finite_check"].all()
    thermo = thermo_physical_constraints_dataframe()
    assert not thermo.empty and thermo["passed"].all()
    transport = transport_physical_constraints_dataframe()
    assert not transport.empty and transport["passed"].all()


def test_v5_3_report_and_repro_package_include_math_kernel_metadata():
    result = run_flowsheet(load_default_config())
    excel = export_excel(result)
    workbook = load_workbook(io.BytesIO(excel), read_only=True)
    required = {
        "dimensioned_inputs",
        "residual_system_detailed",
        "equation_binding",
        "rhs_diagnostics",
        "thermo_physical_constraints",
        "transport_physical_constraints",
        "fallback_diagnostics",
    }
    assert required.issubset(set(workbook.sheetnames))
    package = export_repro_package(result)
    manifest = load_repro_manifest_from_zip(package)
    assert manifest["app_version"].startswith("V6.4")
    with zipfile.ZipFile(io.BytesIO(package)) as zf:
        names = set(zf.namelist())
    assert {"residual_system.csv", "benchmark_snapshot.json", "model_registry_snapshot.json", "equation_registry_snapshot.json"}.issubset(names)


def test_v5_3_doe_and_windows_respect_residual_constraints():
    cfg = load_default_config()
    doe = recommend_next_experiment_batch(cfg, n=3, seed=13)
    assert not doe.empty
    residual_cols = [col for col in doe.columns if col.startswith("feasible_residual_system")]
    assert residual_cols and doe[residual_cols[0]].all()
    windows = generate_feasible_windows(cfg)
    assert windows
    assert all(window.robustness_score >= 0.0 for window in windows)

