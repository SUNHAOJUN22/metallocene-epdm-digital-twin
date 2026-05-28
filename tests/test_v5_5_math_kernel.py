"""V5.5 residual-solver, RHS coupling and benchmark-calibration tests."""

from __future__ import annotations

from io import BytesIO
import zipfile

import numpy as np
import pandas as pd
from openpyxl import load_workbook

from epdm_sim.benchmark_calibration import (
    benchmark_calibration_summary,
    benchmark_residual_dataframe,
    benchmark_weight_by_confidence,
    compare_model_to_experimental_benchmark,
    recommend_calibration_data_gaps,
    update_model_confidence_from_benchmarks,
)
from epdm_sim.dynamic_core.residual_timeseries import dynamic_residual_timeseries, dynamic_rhs_residual_acceptance
from epdm_sim.dynamic_core.rhs_terms import rhs_term_schema, rhs_terms_for_state, rhs_terms_from_profile
from epdm_sim.dynamic_template_reactor import simulate_template_semibatch_ode
from epdm_sim.estimation.confidence import confidence_interval_dataframe
from epdm_sim.estimation.constraints import estimation_parameter_constraints
from epdm_sim.estimation.objectives import residual_aware_parameter_objective
from epdm_sim.estimation.persistence import calibrated_set_record
from epdm_sim.estimation.residuals import parameter_residual_dataframe
from epdm_sim.flowsheet import load_default_config, run_flowsheet
from epdm_sim.flowsheet_core.feed_builder import build_feed_from_template_config
from epdm_sim.flowsheet_core.kpi_builder import build_kpis_for_template
from epdm_sim.flowsheet_core.recycle import recycle_closure_correction
from epdm_sim.flowsheet_core.residual_builder import build_flowsheet_residuals
from epdm_sim.flowsheet_core.unit_sequence import default_unit_sequence
from epdm_sim.fluid_core.density import density_kg_m3
from epdm_sim.fluid_core.heat_capacity import mixture_heat_capacity_kJ_kgK
from epdm_sim.fluid_core.hydraulics import darcy_pressure_drop_kPa
from epdm_sim.fluid_core.viscosity import viscosity_Pa_s
from epdm_sim.kinetics import KineticParameters
from epdm_sim.reactor_core.material_balance import consumed_monomer_mass_kg_h
from epdm_sim.reactor_core.rate_engine import template_rate_engine
from epdm_sim.reactor_core.reactor_outputs import reactor_output_dataframe
from epdm_sim.reactor_core.stoichiometry import monomer_segment_map
from epdm_sim.report import export_excel
from epdm_sim.repro_package import export_repro_package, load_repro_manifest_from_zip
from epdm_sim.residual_solver import (
    ResidualCorrection,
    adjust_flash_split_to_close_mass,
    heat_balance_residual_correction,
    residual_acceptance_summary,
    residual_correction_trace_dataframe,
    residual_solver_dataframe,
    residual_weighted_objective,
    solve_recycle_with_residual_minimization,
)
from epdm_sim.residual_system import build_flowsheet_residual_system
from epdm_sim.state_vector import build_state_layout_from_template, pack_state
from epdm_sim.template_config import process_config_to_template_config
from epdm_sim.template_ode_rhs import build_template_ode_context, initial_template_ode_state


def test_residual_solver_rejects_large_nonphysical_corrections():
    cfg = load_default_config()
    result = run_flowsheet(cfg)
    system = build_flowsheet_residual_system(result)
    assert residual_weighted_objective(system) >= 0.0
    summary = residual_acceptance_summary(system)
    assert summary["passed"]
    assert residual_solver_dataframe(system)["passed"].iloc[0]

    recycle = solve_recycle_with_residual_minimization(100.0, 100.1, max_relative_correction_pct=1.0)
    assert recycle.accepted
    flash_small = adjust_flash_split_to_close_mass(100.0, 10.0, 89.9)
    assert flash_small.accepted
    flash_large = adjust_flash_split_to_close_mass(100.0, 10.0, 50.0)
    assert not flash_large.accepted
    assert flash_large.severity == "critical"
    heat_bad = heat_balance_residual_correction(10.0, 1.0)
    assert not heat_bad.accepted
    manual = ResidualCorrection("manual", "target", 1.0, 1.0, 0.0, "kg/h", 0.0, "ok", True, "", "")
    assert manual.as_dict()["accepted"] is True
    trace = residual_correction_trace_dataframe([recycle, flash_small, heat_bad])
    assert len(trace) == 3
    assert trace["relative_correction_pct"].ge(0.0).all()


def test_dynamic_rhs_terms_are_coupled_to_residuals():
    cfg = load_default_config()
    dynamic = simulate_template_semibatch_ode(config=cfg, total_time_min=4.0, dt_min=1.0, solver_mode="explicit_bounded")
    assert not dynamic.profile.empty
    rhs_profile = rhs_terms_from_profile(dynamic)
    assert not rhs_profile.empty
    assert rhs_profile["passed"].all()
    residual_ts = dynamic_residual_timeseries(dynamic)
    assert not residual_ts.empty
    assert residual_ts["rhs_coupled"].all()
    acceptance = dynamic_rhs_residual_acceptance(dynamic)
    assert acceptance["passed"]
    assert rhs_term_schema()["term_id"].str.contains("feed_term").any()

    layout = build_state_layout_from_template()
    context = build_template_ode_context("EPDM_EPM_metallocene_solution", layout, KineticParameters(), cfg, total_time_min=4.0)
    state = initial_template_ode_state(context, solvent_mass_kg=1.0, temperature_K=373.15, pressure_Pa=1.0e6, catalyst_active_mol=1.0e-6)
    y = pack_state(layout, state)
    rhs_terms = rhs_terms_for_state(0.0, y, context)
    assert not rhs_terms.empty
    assert rhs_terms["finite_check"].all()


def test_benchmark_calibration_weights_and_data_gaps_are_finite():
    plant_weight = benchmark_weight_by_confidence("plant", "high")
    synthetic_weight = benchmark_weight_by_confidence("synthetic", "low")
    assert plant_weight > synthetic_weight > 0.0

    outputs = {"polymer_mass_closure_pct": 0.0, "henry_pressure_monotonic": 1.0}
    comparison = compare_model_to_experimental_benchmark(outputs)
    assert not comparison.empty
    assert comparison["weight"].between(0.0, 1.0).all()
    residuals = benchmark_residual_dataframe(outputs)
    assert np.isfinite(residuals["weighted_abs_residual"].to_numpy()).all()
    confidence = update_model_confidence_from_benchmarks(70.0, outputs)
    assert 0.0 <= confidence["adjusted_score"] <= 100.0
    gaps = recommend_calibration_data_gaps(outputs)
    assert not gaps.empty
    assert gaps["recommended_action"].str.len().gt(0).all()
    assert not benchmark_calibration_summary(outputs).empty


def test_split_math_core_helpers_preserve_api_behavior():
    cfg = load_default_config()
    template_cfg = process_config_to_template_config(cfg)
    feed = build_feed_from_template_config(template_cfg)
    assert feed.total_mass_flow() > 0.0
    assert "reactor" in default_unit_sequence()
    assert recycle_closure_correction(10.0, 10.0).accepted

    result = run_flowsheet(cfg)
    assert build_kpis_for_template("EPDM_EPM_metallocene_solution", result)
    assert build_flowsheet_residuals(result).overall_score >= 70.0

    seg_map = monomer_segment_map()
    assert seg_map["ethylene"] == "E"
    rates = template_rate_engine(
        "EPDM_EPM_metallocene_solution",
        {"ethylene": 1.0, "propylene": 1.0, "ENB": 0.1},
        {"Cstar_mol_L": 1.0e-6},
        {"temperature_K": 373.15, "pressure_MPa": 1.0},
        KineticParameters(),
    )
    assert all(value >= 0.0 for value in rates.values())
    assert consumed_monomer_mass_kg_h("EPDM_EPM_metallocene_solution", {"ethylene": 10.0}) > 0.0
    assert not reactor_output_dataframe({"polymer_kg_h": 1.0}).empty

    assert density_kg_m3((650.0, "kg/m3")) > 0.0
    assert viscosity_Pa_s((1.0, "cP")) == 0.001
    assert mixture_heat_capacity_kJ_kgK(2.2, 1.9, 10.0) > 0.0
    assert darcy_pressure_drop_kPa(100.0, 650.0, 0.01, 0.05, 10.0) >= 0.0

    residuals = parameter_residual_dataframe({"Mw": 300000.0}, {"Mw": 310000.0}, {"Mw": "g/mol"})
    assert residuals["residual"].iloc[0] == 10000.0
    assert residual_aware_parameter_objective(1.0, build_flowsheet_residuals(result)) >= 1.0
    assert not estimation_parameter_constraints().empty
    ci = confidence_interval_dataframe(pd.DataFrame({"k": [1.0, 2.0, 3.0]}))
    assert ci["low"].iloc[0] <= ci["high"].iloc[0]
    record = calibrated_set_record("cal_1", {"k": 1.0}, source_dataset_id="ds")
    assert record["status"] == "calibrated_candidate"


def test_report_and_repro_include_v5_5_audit_sheets():
    result = run_flowsheet(load_default_config())
    xlsx = export_excel(result)
    workbook = load_workbook(BytesIO(xlsx), read_only=True)
    required = {
        "residual_solver",
        "residual_correction_trace",
        "rhs_term_diagnostics",
        "benchmark_calibration",
        "benchmark_data_gaps",
        "posterior_residual_acceptance",
        "uncertainty_residual_risk",
    }
    assert required.issubset(set(workbook.sheetnames))

    package = export_repro_package(result, report_xlsx=xlsx, test_status="v5_5_test")
    manifest = load_repro_manifest_from_zip(package)
    assert manifest["app_version"].startswith("V6.4")
    with zipfile.ZipFile(BytesIO(package)) as zf:
        names = set(zf.namelist())
    assert {"benchmark_residuals.csv", "benchmark_data_gaps.csv", "unit_conversion_trace.csv", "residual_correction_trace.csv"}.issubset(names)

