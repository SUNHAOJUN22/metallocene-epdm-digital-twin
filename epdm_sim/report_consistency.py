"""Report/repro-package metadata consistency checks."""

from __future__ import annotations

import json
import zipfile
from dataclasses import dataclass
from io import BytesIO
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import load_workbook


REQUIRED_EXCEL_SHEETS = [
    "stream table",
    "unit operations",
    "product properties",
    "export_metadata",
    "template_config",
    "template_flowsheet",
    "validity_envelope",
    "performance_profile",
    "report_consistency",
    "calibration_scores",
    "residual_system",
    "equation_binding",
    "ode_diagnostics",
    "thermo_math_core",
    "transport_core",
    "parameter_constraints",
    "extrapolation_risk",
    "benchmark_acceptance",
    "dimensioned_inputs",
    "unit_conversion_trace",
    "residual_system_detailed",
    "residual_objective",
    "dynamic_residuals",
    "phase_equilibrium_constraints",
    "experimental_benchmarks",
    "residual_aware_optimization",
    "residual_aware_doe",
    "rhs_diagnostics",
    "thermo_physical_constraints",
    "transport_physical_constraints",
    "fallback_diagnostics",
    "residual_solver",
    "residual_correction_trace",
    "conservation_correction",
    "correction_certificates",
    "rhs_term_diagnostics",
    "benchmark_calibration",
    "benchmark_data_gaps",
    "data_lineage",
    "residual_constrained_fit",
    "posterior_residual_filter",
    "equation_reverse_check",
    "dynamic_residual_feedback",
    "calibrated_property_models",
    "posterior_residual_acceptance",
    "uncertainty_residual_risk",
    "dynamic_residual_timeseries",
    "equation_residual_coupling",
    "residual_acceptance",
    "dynamic_stability_checks",
    "benchmark_sources",
    "benchmark_lineage",
    "calibrated_property_usage",
    "calibration_lineage",
    "model_traceability_graph",
    "equation_graph",
    "residual_graph",
    "data_lineage_graph",
    "solver_certificates",
    "dae_constraints",
    "state_invariants",
    "validation_evidence",
    "model_confidence",
    "confidence_decomposition",
    "property_model_selection",
    "property_model_bridge",
    "property_model_runtime",
    "residual_aware_calibration",
    "residual_aware_posterior",
    "residual_aware_doe",
    "residual_aware_optimizer",
    "dynamic_solver_decision",
    "dynamic_solver_policy",
    "dynamic_step_acceptance",
    "residual_aware_decision",
    "evidence_chain",
    "evidence_gaps",
    "evidence_chain_score",
    "evidence_gap_priority",
    "conservation_solve_path",
    "conservation_solve_cert",
    "equation_oriented_solver",
    "conservation_jacobian",
    "calibration_data_package",
    "data_assimilation",
    "property_runtime_context",
    "adaptive_step_control",
    "dynamic_event_detection",
    "residual_aware_sampling",
    "confidence_certificate",
    "validation_upgrade_plan",
    "nonlinear_residual_loop",
    "solve_path_integrator",
    "industrial_data_package",
    "benchmark_reconciliation",
    "property_runtime_audit",
    "adaptive_integrator",
    "event_localization",
    "residual_decision_engine",
    "governance_certificate",
    "V6_4_audit_summary",
    "V6_2_audit_summary",
    "V6_3_audit_summary",
]

HEAVY_TASKS = {"dynamic_ode", "cfd", "optimization", "posterior", "doe", "uncertainty"}


@dataclass(frozen=True)
class ReportConsistencyResult:
    """One report consistency gate result."""

    check_id: str
    passed: bool
    severity: str
    message: str
    detail: str = ""

    def as_dict(self) -> dict[str, Any]:
        return self.__dict__.copy()


def excel_sheet_names(excel_bytes: bytes) -> list[str]:
    """Return workbook sheet names without writing to disk."""
    workbook = load_workbook(BytesIO(excel_bytes), read_only=True, data_only=True)
    return list(workbook.sheetnames)


def check_excel_required_sheets(excel_bytes: bytes, required: list[str] | None = None) -> list[ReportConsistencyResult]:
    """Check required report sheets are present."""
    required = required or REQUIRED_EXCEL_SHEETS
    names = set(excel_sheet_names(excel_bytes))
    missing = [sheet for sheet in required if sheet not in names]
    return [
        ReportConsistencyResult(
            "excel_required_sheets",
            not missing,
            "ok" if not missing else "error",
            "Excel workbook contains required audit sheets." if not missing else "Excel workbook is missing required audit sheets.",
            ", ".join(missing),
        )
    ]


def read_excel_metadata(excel_bytes: bytes) -> dict[str, Any]:
    """Read the first row of the export_metadata sheet."""
    try:
        df = pd.read_excel(BytesIO(excel_bytes), sheet_name="export_metadata")
    except Exception:
        return {}
    if df.empty:
        return {}
    return {str(key): value for key, value in df.iloc[0].dropna().to_dict().items()}


def load_repro_manifest(package: bytes | str | Path) -> dict[str, Any]:
    """Load manifest.json from a repro package zip."""
    if isinstance(package, (str, Path)):
        handle: Any = Path(package)
    else:
        handle = BytesIO(package)
    try:
        with zipfile.ZipFile(handle) as zf:
            with zf.open("manifest.json") as fh:
                return json.loads(fh.read().decode("utf-8"))
    except Exception:
        return {}


def compare_report_manifest_metadata(excel_bytes: bytes, manifest: dict[str, Any] | None = None) -> list[ReportConsistencyResult]:
    """Compare stable metadata keys between report and repro manifest."""
    metadata = read_excel_metadata(excel_bytes)
    manifest = manifest or {}
    keys = ["version", "config_hash", "parameter_set_id", "template_id", "model_registry_hash", "equation_registry_hash"]
    rows: list[ReportConsistencyResult] = []
    for key in keys:
        metadata_key = "software_version" if key == "version" else key
        report_value = str(metadata.get(metadata_key, ""))
        manifest_value = str(manifest.get("app_version" if key == "version" else key, ""))
        if not manifest:
            rows.append(ReportConsistencyResult(f"metadata_{key}", bool(report_value), "warning", f"Report metadata includes {key}.", report_value))
        else:
            rows.append(
                ReportConsistencyResult(
                    f"metadata_{key}",
                    bool(report_value) and report_value == manifest_value,
                    "ok" if bool(report_value) and report_value == manifest_value else "warning",
                    f"Report/repro metadata comparison for {key}.",
                    f"report={report_value}; manifest={manifest_value}",
                )
            )
    return rows


def check_export_does_not_run_heavy(task_status: pd.DataFrame | dict[str, Any] | None = None) -> list[ReportConsistencyResult]:
    """Check report export metadata does not indicate hidden heavy-task execution."""
    if task_status is None:
        return [ReportConsistencyResult("export_no_hidden_heavy_tasks", True, "ok", "No task status supplied; report export is treated as read-only by contract.", "")]
    df = task_status if isinstance(task_status, pd.DataFrame) else pd.DataFrame.from_dict(task_status, orient="index").reset_index(names="task_id")
    triggered: list[str] = []
    for _, row in df.iterrows():
        task_id = str(row.get("task_id", row.get("index", "")))
        status = str(row.get("status", "")).lower()
        if task_id in HEAVY_TASKS and status in {"running", "success", "completed"}:
            triggered.append(task_id)
    return [
        ReportConsistencyResult(
            "export_no_hidden_heavy_tasks",
            not triggered,
            "ok" if not triggered else "error",
            "Report export must not trigger heavy model tasks.",
            ", ".join(triggered),
        )
    ]


def report_consistency_dataframe(
    excel_bytes: bytes | None = None,
    *,
    repro_manifest: dict[str, Any] | None = None,
    task_status: pd.DataFrame | dict[str, Any] | None = None,
) -> pd.DataFrame:
    """Return all report consistency checks as a table."""
    results: list[ReportConsistencyResult] = []
    if excel_bytes:
        results.extend(check_excel_required_sheets(excel_bytes))
        results.extend(compare_report_manifest_metadata(excel_bytes, repro_manifest))
    else:
        results.append(ReportConsistencyResult("excel_supplied", False, "warning", "No Excel bytes supplied for report consistency check.", ""))
    results.extend(check_export_does_not_run_heavy(task_status))
    return pd.DataFrame([item.as_dict() for item in results])
